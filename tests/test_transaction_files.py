"""Testes de leitura, exportação e persistência de transações."""

from __future__ import annotations

from io import BytesIO
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.transaction_files import (
    EXCEL_ACCENT_COLOR,
    EXCEL_HEADER_COLOR,
    INSTRUCTIONS_SHEET_NAME,
    TRANSACTION_HEADER_LABELS,
    TRANSACTION_SHEET_NAME,
    create_excel_template,
    export_transactions_to_excel,
    find_matching_transactions,
    normalize_transaction_headers,
    prepare_transactions_for_export,
    read_csv_transactions,
    read_excel_transactions,
    read_ofx_transactions,
    read_transaction_file,
    split_imported_transactions_by_match,
)
from src.transaction_validation import (
    REQUIRED_TRANSACTION_COLUMNS,
    split_transactions_by_validity,
)


def create_test_transactions() -> pd.DataFrame:
    """Cria transações válidas com colunas técnicas adicionais."""
    return pd.DataFrame(
        {
            "data": pd.to_datetime(
                [
                    "2026-08-01",
                    "2026-08-02",
                ]
            ),
            "tipo": [
                "receita",
                "despesa",
            ],
            "descricao": [
                "Bolsa-estágio",
                "Compra no mercado",
            ],
            "categoria": [
                "Trabalho",
                "Alimentação",
            ],
            "valor": [
                1600.00,
                200.50,
            ],
            "arquivo_origem": [
                "transacoes_2026_08.csv",
                "transacoes_2026_08.csv",
            ],
            "ano_mes": [
                "2026-08",
                "2026-08",
            ],
        }
    )


def test_read_csv_transactions_reads_uploaded_content() -> None:
    """Verifica a leitura de um CSV enviado em memória."""
    csv_content = (
        "DATA,TIPO,DESCRIÇÃO,CATEGORIA,VALOR\n"
        "2026-08-01,receita,Bolsa-estágio,Trabalho,1600.00\n"
    )

    source = BytesIO(
        csv_content.encode("utf-8-sig")
    )

    transactions = read_csv_transactions(
        source
    )

    assert len(transactions) == 1

    assert (
        transactions.columns.tolist()
        == REQUIRED_TRANSACTION_COLUMNS
    )

    assert (
        transactions.loc[0, "tipo"]
        == "receita"
    )

    assert (
        transactions.loc[0, "valor"]
        == 1600.00
    )


def test_normalize_transaction_headers_accepts_display_labels() -> None:
    """Converte cabeçalhos visuais para o contrato interno."""
    transactions = pd.DataFrame(
        columns=[
            "DATA",
            "TIPO",
            "DESCRIÇÃO",
            "CATEGORIA",
            "VALOR",
        ]
    )

    normalized = normalize_transaction_headers(
        transactions
    )

    assert (
        normalized.columns.tolist()
        == REQUIRED_TRANSACTION_COLUMNS
    )


def test_prepare_transactions_for_export_removes_technical_columns() -> None:
    """Remove colunas internas antes da exportação."""
    transactions = create_test_transactions()

    result = prepare_transactions_for_export(
        transactions
    )

    assert (
        result.columns.tolist()
        == REQUIRED_TRANSACTION_COLUMNS
    )

    assert (
        "arquivo_origem"
        not in result.columns
    )

    assert (
        "ano_mes"
        not in result.columns
    )

    assert result["data"].tolist() == [
        "2026-08-01",
        "2026-08-02",
    ]


def test_prepare_transactions_for_export_rejects_missing_columns() -> None:
    """Impede exportação quando uma coluna obrigatória está ausente."""
    transactions = pd.DataFrame(
        {
            "data": ["2026-08-01"],
            "tipo": ["receita"],
            "descricao": ["Bolsa-estágio"],
            "valor": [1600.00],
        }
    )

    with pytest.raises(
        ValueError
    ) as error:
        prepare_transactions_for_export(
            transactions
        )

    assert (
        "categoria"
        in str(error.value)
    )


def test_export_transactions_to_excel_can_be_imported_again() -> None:
    """Verifica se um arquivo exportado pode ser lido novamente."""
    transactions = create_test_transactions()

    excel_content = (
        export_transactions_to_excel(
            transactions
        )
    )

    imported_transactions = (
        read_excel_transactions(
            BytesIO(excel_content)
        )
    )

    assert (
        imported_transactions.columns.tolist()
        == REQUIRED_TRANSACTION_COLUMNS
    )

    assert len(
        imported_transactions
    ) == 2

    assert (
        imported_transactions.loc[
            0,
            "descricao",
        ]
        == "Bolsa-estágio"
    )

    assert (
        imported_transactions.loc[
            1,
            "valor",
        ]
        == 200.50
    )
def test_read_ofx_transactions_maps_statement_to_contract() -> None:
    fixture_path = Path(
        "tests/fixtures/ofx/"
        "bank_statement.ofx"
    )

    transactions = (
        read_ofx_transactions(
            fixture_path
        )
    )

    assert (
        transactions.columns.tolist()
        == REQUIRED_TRANSACTION_COLUMNS
    )

    assert len(
        transactions
    ) == 2

    assert transactions.to_dict(
        orient="records"
    ) == [
        {
            "data": "2026-08-01",
            "tipo": "receita",
            "descricao": (
                "Bolsa-estagio"
            ),
            "categoria": (
                "Não categorizado"
            ),
            "valor": 1600.0,
        },
        {
            "data": "2026-08-02",
            "tipo": "despesa",
            "descricao": (
                "Mercado Teste"
            ),
            "categoria": (
                "Não categorizado"
            ),
            "valor": 200.5,
        },
    ]

def test_read_ofx_transactions_supports_credit_card_statement() -> None:
    fixture_path = Path(
        "tests/fixtures/ofx/"
        "credit_card_statement.ofx"
    )

    transactions = read_ofx_transactions(
        fixture_path
    )

    assert (
        transactions.columns.tolist()
        == REQUIRED_TRANSACTION_COLUMNS
    )

    assert len(
        transactions
    ) == 2

    assert transactions.to_dict(
        orient="records"
    ) == [
        {
            "data": "2026-08-05",
            "tipo": "despesa",
            "descricao": (
                "Restaurante Teste"
            ),
            "categoria": (
                "Não categorizado"
            ),
            "valor": 120.9,
        },
        {
            "data": "2026-08-07",
            "tipo": "receita",
            "descricao": (
                "Loja Teste"
            ),
            "categoria": (
                "Não categorizado"
            ),
            "valor": 30.0,
        },
    ]

@pytest.mark.parametrize(
    (
        "removed_fragments",
        "expected_description",
    ),
    [
        (
            (
                b"<NAME>Bolsa-estagio</NAME>",
            ),
            "Pagamento mensal",
        ),
        (
            (
                b"<NAME>Bolsa-estagio</NAME>",
                b"<MEMO>Pagamento mensal</MEMO>",
            ),
            "credit-001",
        ),
    ],
)
def test_read_ofx_transactions_uses_description_fallback(
    removed_fragments: tuple[bytes, ...],
    expected_description: str,
) -> None:
    fixture_path = Path(
        "tests/fixtures/ofx/"
        "bank_statement.ofx"
    )

    content = fixture_path.read_bytes()

    for fragment in removed_fragments:
        content = content.replace(
            fragment,
            b"",
        )

    transactions = read_ofx_transactions(
        BytesIO(
            content
        )
    )

    assert (
        transactions.loc[
            0,
            "descricao",
        ]
        == expected_description
    )
def test_ofx_zero_amount_is_rejected_by_common_validation() -> None:
    fixture_path = Path(
        "tests/fixtures/ofx/"
        "bank_statement.ofx"
    )

    content = fixture_path.read_bytes().replace(
        b"<TRNAMT>-200.50</TRNAMT>",
        b"<TRNAMT>0.00</TRNAMT>",
    )

    imported_transactions = (
        read_ofx_transactions(
            BytesIO(
                content
            )
        )
    )

    (
        valid_transactions,
        rejected_transactions,
    ) = split_transactions_by_validity(
        imported_transactions
    )

    assert (
        valid_transactions[
            "descricao"
        ].tolist()
        == [
            "Bolsa-estagio",
        ]
    )

    assert len(
        rejected_transactions
    ) == 1

@pytest.mark.parametrize(
    (
        "content",
        "expected_message",
    ),
    [
        (
            b"",
            "Não foi possível interpretar",
        ),
        (
            b"isto nao e um arquivo ofx",
            "Não foi possível interpretar",
        ),
        (
            (
                b"OFXHEADER:100\n"
                b"DATA:OFXSGML\n"
                b"VERSION:102\n"
                b"\n"
                b"<OFX><STMTTRN>"
            ),
            "não contém um extrato reconhecido",
        ),
    ],
)
def test_read_ofx_transactions_rejects_invalid_content(
    content: bytes,
    expected_message: str,
) -> None:
    with pytest.raises(
        ValueError,
        match=expected_message,
    ):
        read_ofx_transactions(
            BytesIO(
                content
            )
        )
def test_read_ofx_transactions_supports_xml_format() -> None:
    fixture_path = Path(
        "tests/fixtures/ofx/"
        "xml_bank_statement.ofx"
    )

    transactions = read_ofx_transactions(
        fixture_path
    )

    assert (
        transactions.columns.tolist()
        == REQUIRED_TRANSACTION_COLUMNS
    )

    assert transactions.to_dict(
        orient="records"
    ) == [
        {
            "data": "2026-08-10",
            "tipo": "receita",
            "descricao": (
                "Pagamento XML"
            ),
            "categoria": (
                "Não categorizado"
            ),
            "valor": 900.0,
        },
        {
            "data": "2026-08-11",
            "tipo": "despesa",
            "descricao": (
                "Compra XML"
            ),
            "categoria": (
                "Não categorizado"
            ),
            "valor": 75.25,
        },
    ]

def test_read_transaction_file_dispatches_csv() -> None:
    csv_content = (
        "DATA,TIPO,DESCRIÇÃO,CATEGORIA,VALOR\n"
        "2026-08-01,receita,Bolsa-estágio,Trabalho,1600.00\n"
    )

    transactions = read_transaction_file(
        source=BytesIO(
            csv_content.encode(
                "utf-8-sig"
            )
        ),
        file_name="transacoes.csv",
    )

    assert len(
        transactions
    ) == 1

    assert (
        transactions.columns.tolist()
        == REQUIRED_TRANSACTION_COLUMNS
    )

    assert (
        transactions.loc[
            0,
            "descricao",
        ]
        == "Bolsa-estágio"
    )


def test_read_transaction_file_dispatches_excel() -> None:
    excel_content = (
        export_transactions_to_excel(
            create_test_transactions()
        )
    )

    transactions = read_transaction_file(
        source=BytesIO(
            excel_content
        ),
        file_name="transacoes.xlsx",
    )

    assert len(
        transactions
    ) == 2

    assert (
        transactions.columns.tolist()
        == REQUIRED_TRANSACTION_COLUMNS
    )

def test_read_transaction_file_dispatches_ofx() -> None:
    fixture_path = Path(
        "tests/fixtures/ofx/"
        "bank_statement.ofx"
    )

    with fixture_path.open(
        "rb"
    ) as source:
        transactions = read_transaction_file(
            source=source,
            file_name="bank_statement.ofx",
        )

    assert (
        transactions.columns.tolist()
        == REQUIRED_TRANSACTION_COLUMNS
    )

    assert len(
        transactions
    ) == 2

    assert transactions[
        "tipo"
    ].tolist() == [
        "receita",
        "despesa",
    ]

    assert transactions[
        "valor"
    ].tolist() == [
        1600.0,
        200.5,
    ]

def test_read_transaction_file_explains_missing_excel_sheet() -> None:
    excel_content = BytesIO()

    with pd.ExcelWriter(
        excel_content,
        engine="openpyxl",
    ) as writer:
        pd.DataFrame(
            {
                "valor": [
                    100.0,
                ],
            }
        ).to_excel(
            writer,
            sheet_name="OutraAba",
            index=False,
        )

    with pytest.raises(
        ValueError,
        match=(
            "aba chamada "
            "'Transacoes'"
        ),
    ):
        read_transaction_file(
            source=BytesIO(
                excel_content.getvalue()
            ),
            file_name="transacoes.xlsx",
        )


def test_read_transaction_file_rejects_unsupported_format() -> None:
    with pytest.raises(
        ValueError,
        match="Formato não suportado",
    ):
        read_transaction_file(
            source=BytesIO(
                b"unsupported"
            ),
            file_name="transacoes.pdf",
        )

def test_create_excel_template_contains_expected_sheets() -> None:
    """Verifica as abas existentes no modelo Excel."""
    template_content = create_excel_template()

    excel_file = pd.ExcelFile(
        BytesIO(template_content),
        engine="openpyxl",
    )

    assert excel_file.sheet_names == [
        TRANSACTION_SHEET_NAME,
        INSTRUCTIONS_SHEET_NAME,
    ]


def test_create_excel_template_has_transaction_contract_columns() -> None:
    """Verifica os rótulos visuais e sua normalização."""
    template_content = create_excel_template()

    raw_template = pd.read_excel(
        BytesIO(template_content),
        sheet_name=TRANSACTION_SHEET_NAME,
        engine="openpyxl",
    )

    assert raw_template.empty

    assert (
        raw_template.columns.tolist()
        == list(
            TRANSACTION_HEADER_LABELS.values()
        )
    )

    normalized_template = (
        read_excel_transactions(
            BytesIO(template_content)
        )
    )

    assert (
        normalized_template.columns.tolist()
        == REQUIRED_TRANSACTION_COLUMNS
    )


def test_create_excel_template_contains_instructions() -> None:
    """Verifica o conteúdo básico da aba de instruções."""
    template_content = create_excel_template()

    instructions = pd.read_excel(
        BytesIO(template_content),
        sheet_name=INSTRUCTIONS_SHEET_NAME,
        engine="openpyxl",
    )

    assert len(instructions) == 5

    assert instructions.columns.tolist() == [
        "CAMPO",
        "ORIENTAÇÃO",
        "EXEMPLO",
    ]

    assert instructions["CAMPO"].tolist() == (
        REQUIRED_TRANSACTION_COLUMNS
    )


def test_find_matching_transactions_detects_existing_rows() -> None:
    """Identifica linhas importadas iguais às existentes."""
    imported_transactions = (
        create_test_transactions()
    )

    existing_transactions = (
        create_test_transactions()
        .head(1)
        .copy()
    )

    matches = find_matching_transactions(
        imported_transactions,
        existing_transactions,
    )

    assert len(matches) == 1

    assert (
        matches.iloc[0]["descricao"]
        == "Bolsa-estágio"
    )


def test_split_matching_transactions_respects_occurrence_count() -> None:
    """Respeita a quantidade de ocorrências existentes."""
    base_transaction = (
        create_test_transactions()
        .head(1)
        .copy()
    )

    imported_transactions = pd.concat(
        [
            base_transaction,
            base_transaction,
        ],
        ignore_index=True,
    )

    existing_transactions = (
        base_transaction.copy()
    )

    (
        new_transactions,
        matching_transactions,
    ) = split_imported_transactions_by_match(
        imported_transactions,
        existing_transactions,
    )

    assert len(
        matching_transactions
    ) == 1

    assert len(
        new_transactions
    ) == 1


def test_excel_template_has_visual_configuration() -> None:
    """Verifica estilos, alinhamentos e validações do modelo."""
    template_content = create_excel_template()

    workbook = load_workbook(
        BytesIO(template_content)
    )

    worksheet = workbook[
        TRANSACTION_SHEET_NAME
    ]

    assert (
        worksheet.freeze_panes
        == "A2"
    )

    assert (
        worksheet.auto_filter.ref
        == "A1:E1000"
    )

    assert (
        worksheet.sheet_view.showGridLines
        is True
    )

    assert (
        worksheet["A1"].value
        == "DATA"
    )

    assert (
        worksheet["C1"].value
        == "DESCRIÇÃO"
    )

    assert (
        worksheet["A1"].fill.fill_type
        == "solid"
    )

    assert (
        worksheet["A1"].font.bold
        is True
    )

    assert (
        worksheet["A1"]
        .fill
        .fgColor
        .rgb
        .endswith(EXCEL_HEADER_COLOR)
    )

    assert (
        worksheet["A1"]
        .border
        .bottom
        .color
        .rgb
        .endswith(EXCEL_ACCENT_COLOR)
    )

    assert (
        worksheet.column_dimensions["C"].width
        == 38
    )

    validations = (
        worksheet
        .data_validations
        .dataValidation
    )

    assert len(validations) == 3

    validation_types = {
        validation.type
        for validation in validations
    }

    assert validation_types == {
        "date",
        "list",
        "decimal",
    }

    assert (
        worksheet["A2"].number_format
        == "dd/mm/yyyy"
    )

    assert (
        worksheet["E2"].number_format
        == 'R$ #,##0.00'
    )

    assert (
        worksheet["A2"]
        .alignment
        .horizontal
        == "center"
    )

    assert (
        worksheet["B2"]
        .alignment
        .horizontal
        == "center"
    )

    assert (
        worksheet["C2"]
        .alignment
        .horizontal
        == "left"
    )

    assert (
        worksheet["D2"]
        .alignment
        .horizontal
        == "left"
    )

    assert (
        worksheet["E2"]
        .alignment
        .horizontal
        == "right"
    )


def test_exported_excel_has_table_and_number_formats() -> None:
    """Verifica tabela, cabeçalhos e formatos da exportação."""
    excel_content = export_transactions_to_excel(
        create_test_transactions()
    )

    workbook = load_workbook(
        BytesIO(excel_content)
    )

    worksheet = workbook[
        TRANSACTION_SHEET_NAME
    ]

    assert (
        worksheet.freeze_panes
        == "A2"
    )

    assert (
        worksheet.auto_filter.ref
        == "A1:E3"
    )

    assert len(
        worksheet.tables
    ) == 1

    assert (
        worksheet["A1"].value
        == "DATA"
    )

    assert (
        worksheet["E1"].value
        == "VALOR"
    )

    assert (
        worksheet["A2"].number_format
        == "dd/mm/yyyy"
    )

    assert (
        worksheet["E2"].number_format
        == 'R$ #,##0.00'
    )

    assert (
        worksheet["A2"]
        .alignment
        .horizontal
        == "center"
    )

    assert (
        worksheet["B2"]
        .alignment
        .horizontal
        == "center"
    )

    assert (
        worksheet["C2"]
        .alignment
        .horizontal
        == "left"
    )

    assert (
        worksheet["D2"]
        .alignment
        .horizontal
        == "left"
    )

    assert (
        worksheet["E2"]
        .alignment
        .horizontal
        == "right"
    )