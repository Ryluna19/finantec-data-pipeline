"""Leitura, geração e persistência de arquivos de transações."""

from __future__ import annotations

from codecs import BOM_UTF8
from collections import Counter
from collections.abc import Iterable
from io import BytesIO
from pathlib import Path
from typing import BinaryIO

import pandas as pd
import unicodedata

from ofxparse import OfxParser
from ofxparse.ofxparse import (
    OfxParserException,
)

from src.transaction_identity import (
    TRANSACTION_ID_COLUMN,
    ensure_transaction_ids,
)

from src.transaction_validation import (
    REQUIRED_TRANSACTION_COLUMNS,
    prepare_transactions,
)

from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import (
    Table,
    TableStyleInfo,
)

TRANSACTION_SHEET_NAME = "Transacoes"
INSTRUCTIONS_SHEET_NAME = "Instrucoes"

TRANSACTION_HEADER_LABELS = {
    "data": "DATA",
    "tipo": "TIPO",
    "descricao": "DESCRIÇÃO",
    "categoria": "CATEGORIA",
    "valor": "VALOR",
}

EXCEL_HEADER_COLOR = "1F2937"
EXCEL_HEADER_TEXT_COLOR = "FFFFFF"
EXCEL_ACCENT_COLOR = "F97316"
EXCEL_BORDER_COLOR = "D1D5DB"

EXCEL_TRANSACTION_COLUMN_WIDTHS = {
    "A": 16,
    "B": 14,
    "C": 38,
    "D": 24,
    "E": 18,
}

EXCEL_INSTRUCTION_COLUMN_WIDTHS = {
    "A": 18,
    "B": 48,
    "C": 30,
}

FileSource = str | Path | BinaryIO

STORED_TRANSACTION_COLUMNS = [
    TRANSACTION_ID_COLUMN,
    *REQUIRED_TRANSACTION_COLUMNS,
]

TRANSACTION_COLUMN_ALIASES = {
    "data": (
        "data",
        "data_movimento",
        "data_movimentacao",
        "data_lancamento",
        "data_transacao",
        "dt_movimento",
        "dt_lancamento",
        "date",
    ),
    "tipo": (
        "tipo",
        "tipo_transacao",
        "tipo_movimento",
        "natureza",
        "debito_credito",
        "credito_debito",
    ),
    "descricao": (
        "descricao",
        "historico",
        "historico_lancamento",
        "descricao_lancamento",
        "lancamento",
        "estabelecimento",
        "detalhes",
        "memo",
    ),
    "categoria": (
        "categoria",
        "categoria_transacao",
        "categoria_lancamento",
        "classificacao",
    ),
    "valor": (
        "valor",
        "valor_movimentado",
        "valor_movimento",
        "valor_lancamento",
        "valor_transacao",
        "montante",
        "amount",
    ),
}

DEBIT_COLUMN_ALIASES = (
    "debito",
    "valor_debito",
    "debit",
    "withdrawal",
    "saida",
    "saidas",
)

CREDIT_COLUMN_ALIASES = (
    "credito",
    "valor_credito",
    "credit",
    "deposit",
    "entrada",
    "entradas",
)


def _rewind_file(source: FileSource) -> None:
    """Reposiciona arquivos em memória antes de uma nova leitura."""
    seek = getattr(source, "seek", None)

    if callable(seek):
        seek(0)


def normalize_transaction_column_name(
    column: object,
) -> str:
    """Normaliza um nome de coluna para comparação interna."""
    normalized_column = unicodedata.normalize(
        "NFKD",
        str(column),
    )

    normalized_column = normalized_column.encode(
        "ascii",
        "ignore",
    ).decode(
        "ascii"
    )

    return (
        normalized_column
        .strip()
        .lower()
        .replace(
            " ",
            "_",
        )
    )


def normalize_transaction_headers(
    transactions: pd.DataFrame,
) -> pd.DataFrame:
    """Normaliza os cabeçalhos recebidos para o contrato interno."""
    normalized_transactions = transactions.copy()

    normalized_transactions.columns = [
        normalize_transaction_column_name(
            column
        )
        for column in normalized_transactions.columns
    ]

    return normalized_transactions


def read_csv_transactions(
    source: FileSource,
) -> pd.DataFrame:
    """Lê e normaliza transações de um arquivo CSV."""
    _rewind_file(source)

    transactions = pd.read_csv(
        source,
        encoding="utf-8-sig",
    )

    return normalize_transaction_headers(transactions)


def read_excel_transactions(
    source: FileSource,
    sheet_name: str = TRANSACTION_SHEET_NAME,
) -> pd.DataFrame:
    """Lê e normaliza a planilha principal de um arquivo Excel."""
    _rewind_file(source)

    transactions = pd.read_excel(
        source,
        sheet_name=sheet_name,
        engine="openpyxl",
    )

    return normalize_transaction_headers(transactions)

def read_ofx_transactions(
    source: FileSource,
) -> pd.DataFrame:
    """Converte um extrato OFX para o contrato interno."""
    if isinstance(
        source,
        (
            str,
            Path,
        ),
    ):
        source_content = Path(
            source
        ).read_bytes()

    else:
        _rewind_file(
            source
        )

        source_content = source.read()

    if isinstance(
        source_content,
        str,
    ):
        source_content = (
            source_content.encode(
                "utf-8"
            )
        )

    if source_content.startswith(
        BOM_UTF8
    ):
        source_content = source_content[
            len(BOM_UTF8):
        ]

    try:
        ofx = OfxParser.parse(
            BytesIO(
                source_content
            )
        )

    except (
        OfxParserException,
        ValueError,
    ) as error:
        raise ValueError(
            "Não foi possível interpretar o arquivo OFX. "
            "Verifique se o arquivo é válido e tente novamente."
        ) from error

    account = getattr(
        ofx,
        "account",
        None,
    )

    statement = getattr(
        account,
        "statement",
        None,
    )

    if statement is None:
        raise ValueError(
            "O arquivo OFX não contém "
            "um extrato reconhecido."
        )

    imported_rows: list[
        dict[str, object]
    ] = []

    for transaction in (
        statement.transactions
    ):
        transaction_date = getattr(
            transaction,
            "date",
            None,
        )

        if transaction_date is None:
            raise ValueError(
                "O arquivo OFX possui "
                "uma transação sem data."
            )

        raw_amount = getattr(
            transaction,
            "amount",
            None,
        )

        if raw_amount is None:
            raise ValueError(
                "O arquivo OFX possui "
                "uma transação sem valor."
            )

        amount = float(
            raw_amount
        )

        description = next(
            (
                str(value).strip()
                for value in (
                    getattr(
                        transaction,
                        "payee",
                        None,
                    ),
                    getattr(
                        transaction,
                        "memo",
                        None,
                    ),
                    getattr(
                        transaction,
                        "id",
                        None,
                    ),
                )
                if (
                    value is not None
                    and str(value).strip()
                )
            ),
            "Transação OFX",
        )

        imported_rows.append(
            {
                "data": (
                    transaction_date
                    .date()
                    .isoformat()
                ),
                "tipo": (
                    "receita"
                    if amount > 0
                    else "despesa"
                ),
                "descricao": description,
                "categoria": (
                    "Não categorizado"
                ),
                "valor": abs(
                    amount
                ),
            }
        )

    return pd.DataFrame(
        imported_rows,
        columns=(
            REQUIRED_TRANSACTION_COLUMNS
        ),
    )
    
def list_excel_sheet_names(
    source: FileSource,
) -> list[str]:
    """Lista as abas disponíveis em um arquivo Excel."""
    _rewind_file(
        source
    )

    with pd.ExcelFile(
        source,
        engine="openpyxl",
    ) as workbook:
        return list(
            workbook.sheet_names
        )


def read_excel_table(
    source: FileSource,
    *,
    sheet_name: str,
    header_row: int = 0,
) -> pd.DataFrame:
    """Lê uma tabela Excel preservando os cabeçalhos originais."""
    if (
        not isinstance(
            header_row,
            int,
        )
        or isinstance(
            header_row,
            bool,
        )
        or header_row < 0
    ):
        raise ValueError(
            "A linha de cabeçalho deve ser "
            "um inteiro maior ou igual a zero."
        )

    _rewind_file(
        source
    )

    return pd.read_excel(
        source,
        sheet_name=sheet_name,
        header=header_row,
        engine="openpyxl",
    )

def suggest_transaction_column_mapping(
    columns: Iterable[object],
) -> dict[str, str | None]:
    """Sugere quais colunas externas representam o contrato interno."""
    source_columns: dict[
        str,
        str,
    ] = {}

    for column in columns:
        original_name = str(
            column
        ).strip()

        normalized_name = (
            normalize_transaction_column_name(
                column
            )
        )

        if (
            normalized_name
            and normalized_name
            not in source_columns
        ):
            source_columns[
                normalized_name
            ] = original_name

    suggestions: dict[
        str,
        str | None,
    ] = {}

    for (
        target_column,
        aliases,
    ) in TRANSACTION_COLUMN_ALIASES.items():
        suggestions[
            target_column
        ] = next(
            (
                source_columns[
                    alias
                ]
                for alias in aliases
                if alias in source_columns
            ),
            None,
        )

    return suggestions

def suggest_split_amount_column_mapping(
    columns: Iterable[object],
) -> dict[str, str | None]:
    """Sugere colunas externas separadas de débito e crédito."""
    source_columns: dict[str, str] = {}

    for column in columns:
        original_name = str(
            column
        ).strip()

        normalized_name = (
            normalize_transaction_column_name(
                column
            )
        )

        if (
            normalized_name
            and normalized_name
            not in source_columns
        ):
            source_columns[
                normalized_name
            ] = original_name

    debit_column = next(
        (
            source_columns[alias]
            for alias in DEBIT_COLUMN_ALIASES
            if alias in source_columns
        ),
        None,
    )

    credit_column = next(
        (
            source_columns[alias]
            for alias in CREDIT_COLUMN_ALIASES
            if alias in source_columns
        ),
        None,
    )

    return {
        "debito": debit_column,
        "credito": credit_column,
    }

def suggest_excel_header_row(
    source: FileSource,
    *,
    sheet_name: str,
    max_rows: int = 25,
) -> int:
    """Sugere a linha que contém o cabeçalho da tabela."""
    if (
        not isinstance(max_rows, int)
        or isinstance(max_rows, bool)
        or max_rows <= 0
    ):
        raise ValueError(
            "A quantidade máxima de linhas deve ser "
            "um inteiro maior que zero."
        )

    _rewind_file(source)

    raw_rows = pd.read_excel(
        source,
        sheet_name=sheet_name,
        header=None,
        nrows=max_rows,
        engine="openpyxl",
    )

    if raw_rows.empty:
        return 0

    best_row_index = 0
    best_score = (
        -1,
        -1,
    )

    for row_index, row in raw_rows.iterrows():
        possible_columns = [
            value
            for value in row.tolist()
            if pd.notna(value)
        ]

        suggested_mapping = (
            suggest_transaction_column_mapping(
                possible_columns
            )
        )

        split_amount_mapping = (
            suggest_split_amount_column_mapping(
                possible_columns
            )
        )

        has_single_amount = (
            suggested_mapping[
                "valor"
            ]
            is not None
        )

        has_split_amounts = (
            split_amount_mapping[
                "debito"
            ]
            is not None
            and split_amount_mapping[
                "credito"
            ]
            is not None
        )

        required_score = sum(
            suggested_mapping[field]
            is not None
            for field in (
                "data",
                "descricao",
            )
        )

        required_score += int(
            has_single_amount
            or has_split_amounts
        )

        total_score = sum(
            mapped_column is not None
            for mapped_column
            in suggested_mapping.values()
        )

        total_score += sum(
            mapped_column is not None
            for mapped_column
            in split_amount_mapping.values()
        )

        current_score = (
            required_score,
            total_score,
        )

        if current_score > best_score:
            best_score = current_score
            best_row_index = int(
                row_index
            )

    return best_row_index

def translate_transaction_table(
    transactions: pd.DataFrame,
    column_mapping: dict[
        str,
        str | None,
    ],
    *,
    default_category: str = "Não categorizado",
) -> pd.DataFrame:
    """Traduz uma tabela externa para o contrato interno."""
    if transactions.empty:
        return pd.DataFrame(
            columns=(
                REQUIRED_TRANSACTION_COLUMNS
            )
        )

    required_mappings = (
        "data",
        "descricao",
        "valor",
    )

    missing_mappings = [
        target_column
        for target_column in required_mappings
        if not column_mapping.get(
            target_column
        )
    ]

    if missing_mappings:
        raise ValueError(
            "O mapeamento precisa informar "
            "as colunas de: "
            + ", ".join(
                missing_mappings
            )
            + "."
        )

    mapped_source_columns = {
        target_column: source_column
        for (
            target_column,
            source_column,
        ) in column_mapping.items()
        if source_column
    }

    missing_source_columns = [
        source_column
        for source_column in (
            mapped_source_columns.values()
        )
        if source_column not in transactions.columns
    ]

    if missing_source_columns:
        raise ValueError(
            "As seguintes colunas não foram "
            "encontradas na tabela: "
            + ", ".join(
                sorted(
                    set(
                        missing_source_columns
                    )
                )
            )
            + "."
        )

    normalized_default_category = str(
        default_category
    ).strip()

    if not normalized_default_category:
        raise ValueError(
            "A categoria padrão não pode "
            "ficar vazia."
        )

    data_source = mapped_source_columns[
        "data"
    ]

    description_source = (
        mapped_source_columns[
            "descricao"
        ]
    )

    amount_source = (
        mapped_source_columns[
            "valor"
        ]
    )

    numeric_amounts = pd.to_numeric(
        transactions[
            amount_source
        ],
        errors="coerce",
    )

    type_source = (
        mapped_source_columns.get(
            "tipo"
        )
    )

    if type_source is not None:
        transaction_types = (
            transactions[
                type_source
            ]
            .copy()
        )

    else:
        transaction_types = pd.Series(
            "despesa",
            index=transactions.index,
            dtype="object",
        )

        transaction_types.loc[
            numeric_amounts > 0
        ] = "receita"

    category_source = (
        mapped_source_columns.get(
            "categoria"
        )
    )

    if category_source is not None:
        categories = (
            transactions[
                category_source
            ]
            .copy()
        )

        categories = categories.where(
            categories.notna(),
            normalized_default_category,
        )

        categories = (
            categories
            .astype(
                str
            )
            .str
            .strip()
        )

        categories = categories.mask(
            categories.eq(
                ""
            ),
            normalized_default_category,
        )

    else:
        categories = pd.Series(
            normalized_default_category,
            index=transactions.index,
            dtype="object",
        )

    translated_transactions = (
        pd.DataFrame(
            {
                "data": (
                    transactions[
                        data_source
                    ]
                    .copy()
                ),
                "tipo": transaction_types,
                "descricao": (
                    transactions[
                        description_source
                    ]
                    .copy()
                ),
                "categoria": categories,
                "valor": (
                    numeric_amounts.abs()
                ),
            }
        )
    )

    return (
        translated_transactions[
            REQUIRED_TRANSACTION_COLUMNS
        ]
        .copy()
        .reset_index(
            drop=True
        )
    )

def translate_split_amount_transaction_table(
    transactions: pd.DataFrame,
    *,
    date_column: str,
    description_column: str,
    debit_column: str,
    credit_column: str,
    category_column: str | None = None,
    default_category: str = "Não categorizado",
) -> pd.DataFrame:
    """Traduz débito e crédito separados para o contrato interno."""
    if transactions.empty:
        return pd.DataFrame(
            columns=(
                REQUIRED_TRANSACTION_COLUMNS
            )
        )

    if debit_column == credit_column:
        raise ValueError(
            "As colunas de débito e crédito "
            "precisam ser diferentes."
        )

    source_columns = [
        date_column,
        description_column,
        debit_column,
        credit_column,
    ]

    if category_column is not None:
        source_columns.append(
            category_column
        )

    missing_source_columns = [
        source_column
        for source_column in source_columns
        if source_column
        not in transactions.columns
    ]

    if missing_source_columns:
        raise ValueError(
            "As seguintes colunas não foram "
            "encontradas na tabela: "
            + ", ".join(
                sorted(
                    set(
                        missing_source_columns
                    )
                )
            )
            + "."
        )

    normalized_default_category = str(
        default_category
    ).strip()

    if not normalized_default_category:
        raise ValueError(
            "A categoria padrão não pode "
            "ficar vazia."
        )

    debit_amounts = pd.to_numeric(
        transactions[
            debit_column
        ],
        errors="coerce",
    )

    credit_amounts = pd.to_numeric(
        transactions[
            credit_column
        ],
        errors="coerce",
    )

    has_debit = (
        debit_amounts.notna()
        & debit_amounts.ne(
            0
        )
    )

    has_credit = (
        credit_amounts.notna()
        & credit_amounts.ne(
            0
        )
    )

    valid_debit_rows = (
        has_debit
        & ~has_credit
    )

    valid_credit_rows = (
        has_credit
        & ~has_debit
    )

    transaction_types = pd.Series(
        pd.NA,
        index=transactions.index,
        dtype="object",
    )

    transaction_types.loc[
        valid_debit_rows
    ] = "despesa"

    transaction_types.loc[
        valid_credit_rows
    ] = "receita"

    transaction_amounts = pd.Series(
        float(
            "nan"
        ),
        index=transactions.index,
        dtype="float64",
    )

    transaction_amounts.loc[
        valid_debit_rows
    ] = (
        debit_amounts.loc[
            valid_debit_rows
        ]
        .abs()
    )

    transaction_amounts.loc[
        valid_credit_rows
    ] = (
        credit_amounts.loc[
            valid_credit_rows
        ]
        .abs()
    )

    if category_column is not None:
        categories = (
            transactions[
                category_column
            ]
            .copy()
        )

        categories = categories.where(
            categories.notna(),
            normalized_default_category,
        )

        categories = (
            categories
            .astype(
                str
            )
            .str
            .strip()
        )

        categories = categories.mask(
            categories.eq(
                ""
            ),
            normalized_default_category,
        )

    else:
        categories = pd.Series(
            normalized_default_category,
            index=transactions.index,
            dtype="object",
        )

    translated_transactions = pd.DataFrame(
        {
            "data": (
                transactions[
                    date_column
                ]
                .copy()
            ),
            "tipo": transaction_types,
            "descricao": (
                transactions[
                    description_column
                ]
                .copy()
            ),
            "categoria": categories,
            "valor": transaction_amounts,
        }
    )

    return (
        translated_transactions[
            REQUIRED_TRANSACTION_COLUMNS
        ]
        .copy()
        .reset_index(
            drop=True
        )
    )

def read_transaction_file(
    source: FileSource,
    file_name: str | Path,
) -> pd.DataFrame:
    """Lê transações de acordo com a extensão do arquivo."""
    file_extension = Path(
        file_name
    ).suffix.lower()

    if file_extension == ".csv":
        return read_csv_transactions(
            source
        )

    if file_extension == ".xlsx":
        try:
            return read_excel_transactions(
                source
            )

        except ValueError as error:
            if (
                "Worksheet named"
                in str(
                    error
                )
            ):
                raise ValueError(
                    "O arquivo Excel precisa conter "
                    "uma aba chamada 'Transacoes'."
                ) from error

            raise

    if file_extension == ".ofx":
        return read_ofx_transactions(
            source
        )

    raise ValueError(
        "Formato não suportado. "
        "Envie um arquivo CSV, XLSX ou OFX."
    )

def prepare_transactions_for_export(
    transactions: pd.DataFrame,
) -> pd.DataFrame:
    """Seleciona e formata as colunas destinadas ao usuário."""
    missing_columns = [
        column
        for column in REQUIRED_TRANSACTION_COLUMNS
        if column not in transactions.columns
    ]

    if missing_columns:
        raise ValueError(
            "Não foi possível exportar as transações. "
            "Colunas obrigatórias ausentes: "
            f"{', '.join(missing_columns)}"
        )

    export_data = transactions[REQUIRED_TRANSACTION_COLUMNS].copy()

    export_data["data"] = pd.to_datetime(
        export_data["data"],
        errors="coerce",
    ).dt.strftime("%Y-%m-%d")

    export_data["valor"] = pd.to_numeric(
        export_data["valor"],
        errors="coerce",
    )

    return export_data


def _style_header(
    worksheet,
    column_count: int,
) -> None:
    """Aplica cabeçalho grafite com detalhe inferior laranja."""
    header_fill = PatternFill(
        fill_type="solid",
        fgColor=EXCEL_HEADER_COLOR,
    )

    header_font = Font(
        color=EXCEL_HEADER_TEXT_COLOR,
        bold=True,
    )

    regular_side = Side(
        style="thin",
        color=EXCEL_BORDER_COLOR,
    )

    accent_side = Side(
        style="medium",
        color=EXCEL_ACCENT_COLOR,
    )

    header_border = Border(
        left=regular_side,
        right=regular_side,
        bottom=accent_side,
    )

    for cell in worksheet[1][:column_count]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    worksheet.row_dimensions[1].height = 26


def _set_column_widths(
    worksheet,
    widths: dict[str, int],
) -> None:
    """Define larguras adequadas para as colunas da planilha."""
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width


def _add_transaction_type_validation(
    worksheet,
    last_row: int = 1000,
) -> None:
    """Adiciona uma lista suspensa para o tipo da transação."""
    type_validation = DataValidation(
        type="list",
        formula1='"receita,despesa"',
        allow_blank=True,
    )

    type_validation.error = "Use somente receita ou despesa."

    type_validation.errorTitle = "Tipo de transação inválido"

    type_validation.prompt = "Selecione receita ou despesa."

    type_validation.promptTitle = "Tipo da transação"

    type_validation.showErrorMessage = True
    type_validation.showInputMessage = True

    worksheet.add_data_validation(type_validation)

    type_validation.add(f"B2:B{last_row}")


def _add_transaction_date_validation(
    worksheet,
    last_row: int = 1000,
) -> None:
    """Restringe a coluna de data a valores válidos."""
    date_validation = DataValidation(
        type="date",
        operator="between",
        formula1="DATE(2000,1,1)",
        formula2="DATE(2100,12,31)",
        allow_blank=True,
    )

    date_validation.error = "Informe uma data válida no formato DD/MM/AAAA."

    date_validation.errorTitle = "Data inválida"

    date_validation.prompt = "Informe a data da transação no formato DD/MM/AAAA."

    date_validation.promptTitle = "Data da transação"
    date_validation.showErrorMessage = True
    date_validation.showInputMessage = True

    worksheet.add_data_validation(date_validation)

    date_validation.add(f"A2:A{last_row}")


def _add_transaction_amount_validation(
    worksheet,
    last_row: int = 1000,
) -> None:
    """Restringe a coluna de valor a números positivos."""
    amount_validation = DataValidation(
        type="decimal",
        operator="greaterThan",
        formula1="0",
        allow_blank=True,
    )

    amount_validation.error = "Informe um valor numérico maior que zero."

    amount_validation.errorTitle = "Valor inválido"

    amount_validation.prompt = "Informe somente o valor numérico, sem escrever R$."

    amount_validation.promptTitle = "Valor da transação"
    amount_validation.showErrorMessage = True
    amount_validation.showInputMessage = True

    worksheet.add_data_validation(amount_validation)

    amount_validation.add(f"E2:E{last_row}")


def _format_transaction_columns(
    worksheet,
    last_row: int,
) -> None:
    """Aplica formatos e alinhamentos por tipo de informação."""
    centered_alignment = Alignment(
        horizontal="center",
        vertical="center",
    )

    left_alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=True,
    )

    right_alignment = Alignment(
        horizontal="right",
        vertical="center",
    )

    for row_number in range(
        2,
        last_row + 1,
    ):
        date_cell = worksheet[f"A{row_number}"]
        type_cell = worksheet[f"B{row_number}"]
        description_cell = worksheet[f"C{row_number}"]
        category_cell = worksheet[f"D{row_number}"]
        amount_cell = worksheet[f"E{row_number}"]

        date_cell.number_format = "dd/mm/yyyy"
        amount_cell.number_format = "R$ #,##0.00"

        date_cell.alignment = centered_alignment
        type_cell.alignment = centered_alignment

        description_cell.alignment = left_alignment
        category_cell.alignment = left_alignment

        # Valores à direita facilitam comparar quantias verticalmente.
        amount_cell.alignment = right_alignment

        worksheet.row_dimensions[row_number].height = 22


def _style_transaction_sheet(
    worksheet,
    data_row_count: int,
    enable_input_validation: bool = False,
) -> None:
    """Configura aparência e validações da aba de transações."""
    _style_header(
        worksheet,
        column_count=5,
    )

    _set_column_widths(
        worksheet,
        EXCEL_TRANSACTION_COLUMN_WIDTHS,
    )

    worksheet.freeze_panes = "A2"

    # As linhas de grade ajudam no preenchimento manual.
    worksheet.sheet_view.showGridLines = True

    worksheet.sheet_properties.tabColor = EXCEL_ACCENT_COLOR

    if enable_input_validation:
        last_row = 1000

        worksheet.auto_filter.ref = f"A1:E{last_row}"

        _add_transaction_date_validation(
            worksheet,
            last_row=last_row,
        )

        _add_transaction_type_validation(
            worksheet,
            last_row=last_row,
        )

        _add_transaction_amount_validation(
            worksheet,
            last_row=last_row,
        )

        _format_transaction_columns(
            worksheet,
            last_row=last_row,
        )

        return

    last_row = data_row_count + 1

    if data_row_count <= 0:
        return

    table_reference = f"A1:E{last_row}"

    worksheet.auto_filter.ref = table_reference

    _format_transaction_columns(
        worksheet,
        last_row=last_row,
    )

    transactions_table = Table(
        displayName="FinanTecTransactions",
        ref=table_reference,
    )

    transactions_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    worksheet.add_table(transactions_table)


def _style_instructions_sheet(
    worksheet,
) -> None:
    """Organiza visualmente a aba de instruções."""
    _style_header(
        worksheet,
        column_count=3,
    )

    _set_column_widths(
        worksheet,
        EXCEL_INSTRUCTION_COLUMN_WIDTHS,
    )

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.auto_filter.ref = "A1:C6"

    for row in worksheet.iter_rows(
        min_row=2,
        max_row=6,
        min_col=1,
        max_col=3,
    ):
        for cell in row:
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


def export_transactions_to_excel(
    transactions: pd.DataFrame,
) -> bytes:
    """Gera um arquivo Excel formatado com as transações."""
    export_data = prepare_transactions_for_export(transactions)

    excel_data = export_data.copy()

    # Mantém datas como valores reais do Excel, não apenas texto.
    excel_data["data"] = pd.to_datetime(
        excel_data["data"],
        errors="coerce",
    )
    excel_data = excel_data.rename(columns=TRANSACTION_HEADER_LABELS)

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:
        excel_data.to_excel(
            writer,
            sheet_name=TRANSACTION_SHEET_NAME,
            index=False,
        )

        worksheet = writer.sheets[TRANSACTION_SHEET_NAME]

        _style_transaction_sheet(
            worksheet,
            data_row_count=len(excel_data),
        )

    return output.getvalue()


def create_excel_template() -> bytes:
    """Gera um modelo Excel formatado com instruções."""
    transactions_template = pd.DataFrame(
        columns=[
            TRANSACTION_HEADER_LABELS[column] for column in REQUIRED_TRANSACTION_COLUMNS
        ]
    )

    instructions = pd.DataFrame(
        {
            "CAMPO": [
                "data",
                "tipo",
                "descricao",
                "categoria",
                "valor",
            ],
            "ORIENTAÇÃO": [
                (
                    "Use uma data válida. "
                    "A planilha exibirá DD/MM/AAAA."
                ),
                "Selecione receita ou despesa.",
                "Informe uma descrição curta.",
                "Informe uma categoria financeira.",
                "Use um número positivo, sem R$.",
            ],
            "EXEMPLO": [
                "05/08/2026",
                "despesa",
                "Compra no mercado",
                "Alimentação",
                "200.50",
            ],
        }
    )

    output = BytesIO()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:
        transactions_template.to_excel(
            writer,
            sheet_name=TRANSACTION_SHEET_NAME,
            index=False,
        )

        instructions.to_excel(
            writer,
            sheet_name=INSTRUCTIONS_SHEET_NAME,
            index=False,
        )

        transactions_sheet = writer.sheets[TRANSACTION_SHEET_NAME]

        instructions_sheet = writer.sheets[INSTRUCTIONS_SHEET_NAME]

        _style_transaction_sheet(
            transactions_sheet,
            data_row_count=0,
            enable_input_validation=True,
        )

        _style_instructions_sheet(instructions_sheet)

    return output.getvalue()


def normalize_transaction_keys(
    transactions: pd.DataFrame,
) -> pd.DataFrame:
    """Normaliza os campos usados na comparação de transações."""
    normalized = prepare_transactions(transactions)

    normalized = normalized[REQUIRED_TRANSACTION_COLUMNS].copy()

    normalized["data"] = normalized["data"].dt.strftime("%Y-%m-%d")

    normalized["valor"] = normalized["valor"].round(2)

    return normalized

def split_imported_transactions_by_match(
    imported_transactions: pd.DataFrame,
    existing_transactions: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Separa linhas novas de ocorrências já presentes na base."""
    if imported_transactions.empty:
        empty = imported_transactions.copy()
        return empty, empty

    if existing_transactions.empty:
        return (
            imported_transactions.copy(),
            imported_transactions.iloc[0:0].copy(),
        )

    imported_normalized = normalize_transaction_keys(imported_transactions)

    existing_normalized = normalize_transaction_keys(existing_transactions)

    existing_counts = Counter(
        existing_normalized.itertuples(
            index=False,
            name=None,
        )
    )

    new_positions: list[int] = []
    matching_positions: list[int] = []

    for position, row in enumerate(
        imported_normalized.itertuples(
            index=False,
            name=None,
        )
    ):
        if existing_counts[row] > 0:
            matching_positions.append(position)
            existing_counts[row] -= 1
        else:
            new_positions.append(position)

    new_transactions = imported_transactions.iloc[new_positions].copy()

    matching_transactions = imported_transactions.iloc[matching_positions].copy()

    return (
        new_transactions,
        matching_transactions,
    )


def find_matching_transactions(
    imported_transactions: pd.DataFrame,
    existing_transactions: pd.DataFrame,
) -> pd.DataFrame:
    """Retorna as ocorrências importadas já presentes na base."""
    _, matching_transactions = split_imported_transactions_by_match(
        imported_transactions,
        existing_transactions,
    )

    return matching_transactions
